import streamlit as st
import pandas as pd
import io
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import plotly.express as px
import plotly.graph_objects as go
from st_aggrid import AgGrid, GridOptionsBuilder, GridUpdateMode
import re
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas

# Configurações de Página
st.set_page_config(
    page_title="Dashboard de Projetos",
    page_icon="📊",
    layout="wide"
)

def parse_duracao(duracao_str):
    match = re.search(r'(\d+)', str(duracao_str))
    return int(match.group(1)) if match else None

def parse_date(date_str):
    try:
        return pd.to_datetime(date_str[4:], format='%d/%m/%y', errors='coerce')
    except:
        return pd.NaT

def clean_weekday_abbreviation(date_str):
    return date_str[4:] if isinstance(date_str, str) else date_str

def gerar_curva_s(df_raw, start_date_str='16/09/2024'):
    df_raw['Início'] = df_raw['Início'].apply(lambda x: clean_weekday_abbreviation(x) if isinstance(x, str) else x)
    df_raw['Término'] = df_raw['Término'].apply(lambda x: clean_weekday_abbreviation(x) if isinstance(x, str) else x)

    start_date = pd.to_datetime(start_date_str)
    end_date = df_raw['Término'].max()
    weeks = pd.date_range(start=start_date, end=end_date, freq='W-MON')

    progress_by_week = pd.DataFrame(weeks, columns=['Data'])
    progress_by_week['% Executado'] = 0.0

    for _, row in df_raw.iterrows():
        if pd.notna(row['Início']) and pd.notna(row['Término']):
            task_weeks = pd.date_range(start=row['Início'], end=row['Término'], freq='W-MON')
            weekly_progress = 1 / len(task_weeks) if len(task_weeks) > 0 else 1
            for week in task_weeks:
                progress_by_week.loc[progress_by_week['Data'] == week, '% Executado'] += weekly_progress

    progress_by_week['% Executado Acumulado'] = progress_by_week['% Executado'].cumsum() * 100
    max_progress = progress_by_week['% Executado Acumulado'].max()
    if max_progress > 0:
        progress_by_week['% Executado Acumulado'] = (progress_by_week['% Executado Acumulado'] / max_progress) * 100

    return progress_by_week

def gerar_pdf_cronograma(df):
    pdf_buffer = io.BytesIO()
    c = canvas.Canvas(pdf_buffer, pagesize=letter)
    c.drawString(100, 750, "Relatório de Cronograma")

    y = 730
    for i, (_, row) in enumerate(df.iterrows()):
        try:
            inicio = row['Início'].strftime('%d/%m/%Y') if pd.notna(row['Início']) else "-"
            termino = row['Término'].strftime('%d/%m/%Y') if pd.notna(row['Término']) else "-"
            texto = f"{i + 1}. {row['Nome da tarefa']} | Início: {inicio} | Término: {termino}"
            c.drawString(40, y, texto)
            y -= 20
            if y < 50:
                c.showPage()
                y = 750
        except:
            continue

    c.save()
    pdf_buffer.seek(0)
    return pdf_buffer

# Sidebar
st.sidebar.title("Painel de Controle")
st.sidebar.info("Navegue pelas seções para acessar diferentes dados do projeto.")

uploaded_file = st.file_uploader("Carregar Cronograma", type=["xlsx"])

if uploaded_file is not None:
    df_raw = pd.read_excel(uploaded_file)

    df_raw['Duração'] = df_raw['Duração'].apply(parse_duracao)
    df_raw['Início'] = df_raw['Início'].apply(parse_date)
    df_raw['Término'] = df_raw['Término'].apply(parse_date)

    porcentagem_executada = df_raw.iloc[0]['% concluída'] if '% concluída' in df_raw.columns else 0

    if 'Duração' not in df_raw.columns or 'Início' not in df_raw.columns or 'Término' not in df_raw.columns:
        st.error("O arquivo deve conter as colunas 'Duração', 'Início', e 'Término'.")
    else:
        data_inicio_mais_cedo = df_raw['Início'].min()
        data_termino_mais_tarde = df_raw['Término'].max()
        prazo_total = (data_termino_mais_tarde - data_inicio_mais_cedo).days
        dias_para_finalizar = (data_termino_mais_tarde - pd.Timestamp.today()).days

        atividades_concluidas = len(df_raw[df_raw['% concluída'] == 1])
        atividades_atrasadas = df_raw[(df_raw['Término'] < pd.Timestamp.today()) & (df_raw['% concluída'] != 1)]

        st.title("Dashboard do Projeto")
        col1, col2, col3, col4 = st.columns(4)
        col1.metric("Atividades Concluídas", atividades_concluidas)
        col2.metric("Atividades Atrasadas", len(atividades_atrasadas))
        col3.metric("Prazo Total do Projeto", f"{prazo_total} dias")
        col4.metric("Dias para Finalizar", f"{dias_para_finalizar} dias")

        curva_s_df = gerar_curva_s(df_raw, start_date_str=data_inicio_mais_cedo.strftime('%d/%m/%Y'))
        fig_curva_s = px.line(curva_s_df, x='Data', y='% Executado Acumulado', title="Curva S - Progresso Acumulado do Projeto")
        fig_curva_s.add_trace(go.Scatter(
            x=[pd.Timestamp.today()],
            y=[porcentagem_executada],
            mode="markers+text",
            text=["Este é seu avanço"],
            textposition="top center",
            marker=dict(size=12, color="red")
        ))
        st.plotly_chart(fig_curva_s, use_container_width=True)

        with st.expander("Dados do Cronograma"):
            st.dataframe(df_raw)

        with st.expander("Calendário Interativo de Tarefas"):
            df_calendario = df_raw[['Nome da tarefa', 'Início', 'Término']]
            gd = GridOptionsBuilder.from_dataframe(df_calendario)
            gd.configure_pagination(enabled=True)
            gd.configure_selection(selection_mode="single", use_checkbox=True)
            grid_options = gd.build()
            grid_response = AgGrid(
                df_calendario,
                gridOptions=grid_options,
                update_mode=GridUpdateMode.SELECTION_CHANGED,
                allow_unsafe_jscode=True,
                theme='streamlit',
            )
            selected = grid_response['selected_rows']
            if selected:
                st.write("**Detalhes da Tarefa Selecionada:**")
                st.write(f"Tarefa: {selected[0]['Nome da tarefa']}")
                st.write(f"Início: {selected[0]['Início']}")
                st.write(f"Término: {selected[0]['Término']}")

        if 'Predecessoras' in df_raw.columns:
            atividades_sem_predecessora = df_raw[df_raw['Predecessoras'].isna()]
            with st.expander("Atividades sem Predecessoras"):
                st.dataframe(atividades_sem_predecessora)

        with st.expander("Atividades Atrasadas"):
            st.dataframe(atividades_atrasadas)

        with st.expander("Atividades para Próxima Semana"):
            proxima_semana = pd.Timestamp.today() + pd.Timedelta(days=7)
            atividades_proxima_semana = df_raw[
                (df_raw['Início'] <= proxima_semana) & (df_raw['Término'] >= pd.Timestamp.today())
            ]
            st.dataframe(atividades_proxima_semana)

        with st.expander("Atividades para os Próximos 15 Dias"):
            proximos_15_dias = pd.Timestamp.today() + pd.Timedelta(days=15)
            atividades_proximos_15_dias = df_raw[
                (df_raw['Início'] <= proximos_15_dias) & (df_raw['Término'] >= pd.Timestamp.today())
            ]
            st.dataframe(atividades_proximos_15_dias)

        # Exportar PDF
        pdf_data = gerar_pdf_cronograma(df_raw)
        st.download_button(
            label="📥 Baixar Relatório em PDF",
            data=pdf_data,
            file_name="relatorio_projeto.pdf",
            mime="application/pdf"
        )

        # Exportar Excel
        excel_output = io.BytesIO()
        wb = Workbook()
        ws_curva_s = wb.active
        ws_curva_s.title = 'Curva S'
        for r in dataframe_to_rows(curva_s_df, index=False, header=True):
            ws_curva_s.append(r)

        ws_ativ_semana = wb.create_sheet(title="Atividades Próxima Semana")
        for r in dataframe_to_rows(atividades_proxima_semana, index=False, header=True):
            ws_ativ_semana.append(r)

        ws_ativ_15 = wb.create_sheet(title="Atividades Próximos 15 Dias")
        for r in dataframe_to_rows(atividades_proximos_15_dias, index=False, header=True):
            ws_ativ_15.append(r)

        ws_ativ_atrasadas = wb.create_sheet(title="Atividades Atrasadas")
        for r in dataframe_to_rows(atividades_atrasadas, index=False, header=True):
            ws_ativ_atrasadas.append(r)

        wb.save(excel_output)
        excel_output.seek(0)

        st.download_button(
            label="📥 Baixar Relatório em Excel",
            data=excel_output.getvalue(),
            file_name="relatorio_projeto.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
else:
    st.warning("Por favor, carregue o cronograma para visualizar o painel.")
